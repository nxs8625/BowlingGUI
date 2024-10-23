from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage, messagebox
import openpyxl
import re
import sys



# Determine if the script is running in a PyInstaller bundle
if getattr(sys, "frozen", False):
    OUTPUT_PATH = Path.cwd()
    ASSETS_PATH = Path(sys._MEIPASS) / "assets/frame0"
else:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / "assets/frame0"

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

# Initialize the Tkinter window
window = Tk()
window.geometry("1200x600")
window.configure(bg="#FFFFFF")

canvas = Canvas(
    window,
    bg="#FFFFFF",
    height=600,
    width=1200,
    bd=0,
    highlightthickness=0,
    relief="ridge"
)
canvas.place(x=0, y=0)
canvas.create_rectangle(
    600.0,
    0.0,
    1200.0,
    600.0,
    fill="#747171",
    outline=""
)

image_image_1 = PhotoImage(file=relative_to_assets("image_1.png"))
image_1 = canvas.create_image(300.0, 300.0, image=image_image_1)

canvas.create_text(
    763.0,
    10.0,
    anchor="nw",
    text="Bowling Ball Technical Specifications",
    fill="#FFFFFF",
    font=("Inter ExtraBold", 20 * -1)
)

# Define Entry fields
entry_fields = {
    'First Name': (630.0, 98.0),
    'Last Name': (960.0, 98.0),
    'Ball Name': (630.0, 189.0),
    'Core Type': (960.0, 189.0),
    'RG': (630.0, 266.0),
    'Differential': (960.0, 266.0),
    'Cover Type': (630.0, 377.0),
    'Serial Number': (960.0, 377.0)
}

entries = {}

for label, (x, y) in entry_fields.items():
    entry_image = PhotoImage(file=relative_to_assets(f"entry_{len(entries) + 1}.png"))
    canvas.create_image(
        x + 109.5,
        y + 17.5,
        image=entry_image
    )
    entry = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry.place(
        x=x,
        y=y,
        width=219.0,
        height=25.0
    )
    entries[label] = entry

# Define labels
labels = [
    "First Name", "Last Name", "Ball Name", "Core Type", 
    "RG", "Differential", "Cover Type", "Serial Number"
]

for i, label in enumerate(labels):
    canvas.create_text(
        entry_fields[label][0] + 5,
        entry_fields[label][1] - 25,
        anchor="nw",
        text=label,
        fill="#FFFFFF",
        font=("Inter ExtraBold", 20 * -1)
    )

button_image_1 = PhotoImage(file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: save_to_excel(),
    relief="flat"
)
button_1.place(
    x=750.0,
    y=515.0,
    width=289.0,
    height=71.0
)

button_image_hover_1 = PhotoImage(file=relative_to_assets("button_hover_1.png"))

def button_1_hover(e):
    button_1.config(image=button_image_hover_1)

def button_1_leave(e):
    button_1.config(image=button_image_1)

button_1.bind('<Enter>', button_1_hover)
button_1.bind('<Leave>', button_1_leave)

def save_to_excel():
    # Check if all fields are filled
    data = {label: entry.get() for label, entry in entries.items()}
    
    if any(not value for value in data.values()):
        messagebox.showerror("Input Error", "All fields must be filled.")
        return
    
    # Create the Excel file if it does not exist
    file_path = Path("bowlingballtest.xlsx")
    if file_path.exists():
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        next_id = sheet.max_row
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["ID"] + list(data.keys()))
        next_id = 1
    
    # Save the data to the Excel file
    data_row = [next_id] + list(data.values())
    sheet.append(data_row)
    wb.save(file_path)
    
    # Clear all fields
    for entry in entries.values():
        entry.delete(0, 'end')
    
    # Show success message
    messagebox.showinfo("Success", "Data submitted successfully.")

window.resizable(False, False)
window.mainloop()

